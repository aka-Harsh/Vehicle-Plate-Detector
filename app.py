import cv2
import numpy as np
import pytesseract
import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Set the path to the Tesseract executable (update this path according to your installation)
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # For Windows
# pytesseract.pytesseract.tesseract_cmd = r'/usr/local/bin/tesseract'  # For macOS/Linux

def preprocess_image(image):
    if len(image.shape) == 3:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image
    blur = cv2.GaussianBlur(gray, (5, 5), 0)
    
    # Apply edge detection
    edges = cv2.Canny(blur, 50, 150)
    
    # Dilate the edges to connect them
    kernel = np.ones((3,3), np.uint8)
    dilated = cv2.dilate(edges, kernel, iterations=1)
    
    return dilated

def detect_plate(image):
    contours, _ = cv2.findContours(image, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    contours = sorted(contours, key=cv2.contourArea, reverse=True)[:10]
    
    for contour in contours:
        perimeter = cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, 0.018 * perimeter, True)
        
        if len(approx) == 4:
            x, y, w, h = cv2.boundingRect(contour)
            aspect_ratio = float(w) / h
            
            # Check if the aspect ratio is typical for a license plate
            if 2 <= aspect_ratio <= 5:
                return (x, y, w, h)
    
    return None

def deskew(image):
    coords = np.column_stack(np.where(image > 0))
    angle = cv2.minAreaRect(coords)[-1]
    if angle < -45:
        angle = -(90 + angle)
    else:
        angle = -angle
    (h, w) = image.shape[:2]
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, angle, 1.0)
    rotated = cv2.warpAffine(image, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
    return rotated

def recognize_plate(plate_image):
    # Ensure the image is in grayscale
    if len(plate_image.shape) == 3:
        gray = cv2.cvtColor(plate_image, cv2.COLOR_BGR2GRAY)
    else:
        gray = plate_image
    
    # Apply thresholding
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    
    # Deskew the image
    deskewed = deskew(thresh)
    
    # Apply additional preprocessing
    denoised = cv2.fastNlMeansDenoising(deskewed, None, 10, 7, 21)
    
    # Improve contrast
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    enhanced = clahe.apply(denoised)
    
    # OCR configuration
    config = '--oem 3 --psm 7 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
    
    # Perform OCR
    text = pytesseract.image_to_string(enhanced, config=config)
    text = ''.join(e for e in text if e.isalnum())
    
    return text

def save_to_excel(plate_number, timestamp, plate_image_path, filename='plate_records.xlsx'):
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plate Records"
        headers = ["Plate Number", "Timestamp", "Plate Image"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
    else:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

    # Find the next empty row
    next_row = ws.max_row + 1

    # Insert data
    ws.cell(row=next_row, column=1, value=plate_number)
    ws.cell(row=next_row, column=2, value=timestamp)

    # Insert image
    img = XLImage(plate_image_path)
    img.width = 200  # Adjust as needed
    img.height = 100  # Adjust as needed
    ws.add_image(img, f'C{next_row}')

    # Adjust row height and column width
    ws.row_dimensions[next_row].height = 75  # Adjust as needed
    ws.column_dimensions['C'].width = 30  # Adjust as needed

    # Center align all cells
    for col in range(1, 4):
        cell = ws.cell(row=next_row, column=col)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-fit column width for text columns
    for col in ['A', 'B']:
        ws.column_dimensions[col].auto_size = True

    wb.save(filename)

def main():
    cap = cv2.VideoCapture(0)
    
    if not os.path.exists('plate_images'):
        os.makedirs('plate_images')

    last_plate = ""
    while True:
        ret, frame = cap.read()
        if not ret:
            print("Failed to grab frame")
            break

        processed = preprocess_image(frame)
        plate_coords = detect_plate(processed)

        if plate_coords is not None:
            x, y, w, h = plate_coords
            plate_img = frame[y:y+h, x:x+w]
            plate_number = recognize_plate(plate_img)

            if plate_number and len(plate_number) > 5 and plate_number != last_plate:
                last_plate = plate_number
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                plate_filename = f"plate_images/plate_{plate_number}_{timestamp.replace(':', '-')}.jpg"
                cv2.imwrite(plate_filename, plate_img)

                save_to_excel(plate_number, timestamp, plate_filename)

                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, plate_number, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

        cv2.imshow('Number Plate Detector', frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    main()